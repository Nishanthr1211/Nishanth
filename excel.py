from tkinter import*
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
wd=Tk()
wd.title("Excel Sheet" )
wd.geometry("700x600")

#reate workbook instance
wb=Workbook()
#Load existing workbook
wb=load_workbook("C:\\Users\\Nishanth\\Desktop\\New folder\\sample.xlsr")
#Create active worksheet
ws.wb.active


#Create variable for column A

column_a=ws['A']
column_b=ws['B']

print(column_a)

wd.mainloop()
